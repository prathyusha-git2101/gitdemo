import openpyxl

book = openpyxl.load_workbook("D:\\python excels\\excelhandle.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
Dict = {}
print(cell.value)
sheet.cell(row=2,column=2).value = ("prathyusha")

print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A11'].value)
for i in range(1,sheet.max_row):
     if sheet.cell(row=i,column=1).value == "testcase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)